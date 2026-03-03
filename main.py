"""
StatementExcel - FastAPI Backend
รองรับ KBank, SCB, BBL, KTB
"""

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import uvicorn
import uuid
import os
import asyncio
import time
from pathlib import Path
from typing import Optional

# --- PDF parsing libs ---
# pip install pdfplumber openpyxl pandas camelot-py[cv] fastapi uvicorn python-multipart

import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

app = FastAPI(
    title="StatementExcel API",
    description="แปลง PDF Statement จากธนาคารเป็น Excel",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # production: set to your domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path("tmp/uploads")
OUTPUT_DIR = Path("tmp/outputs")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# =====================
# BANK PARSERS
# =====================

class TransactionRow:
    def __init__(self, date: str, description: str, ref: str,
                 debit: float, credit: float, balance: float, category: str = ""):
        self.date = date
        self.description = description
        self.ref = ref
        self.debit = debit
        self.credit = credit
        self.balance = balance
        self.category = category


def categorize(description: str) -> str:
    """Auto-categorize transactions using keyword matching (Pro feature)."""
    desc = description.lower()
    if any(k in desc for k in ["seven", "7-11", "lotus", "makro", "big c", "tops", "foodland"]):
        return "🛒 ร้านค้า"
    if any(k in desc for k in ["grab", "bolt", "taxi", "bts", "mrt", "แท็กซี่"]):
        return "🚗 เดินทาง"
    if any(k in desc for k in ["kfc", "mcdonald", "pizza", "oishi", "starbucks", "coffee", "bar b q"]):
        return "🍔 อาหาร"
    if any(k in desc for k in ["shopee", "lazada", "amazon", "jd.com"]):
        return "🛍️ ช้อปปิ้งออนไลน์"
    if any(k in desc for k in ["truemove", "ais", "dtac", "true", "internet"]):
        return "📱 โทรศัพท์/อินเทอร์เน็ต"
    if any(k in desc for k in ["transfer", "โอน", "prom pay", "promptpay"]):
        return "💸 โอนเงิน"
    if any(k in desc for k in ["salary", "เงินเดือน", "payroll"]):
        return "💼 เงินเดือน"
    if any(k in desc for k in ["atm", "withdraw", "ถอน"]):
        return "🏧 ถอนเงิน"
    if any(k in desc for k in ["interest", "ดอกเบี้ย"]):
        return "💰 ดอกเบี้ย"
    return "📋 อื่นๆ"


def clean_amount(val: str) -> float:
    """Convert string amount to float."""
    if not val or str(val).strip() in ["", "-", "–"]:
        return 0.0
    cleaned = re.sub(r"[,\s]", "", str(val))
    try:
        return float(cleaned)
    except:
        return 0.0


def parse_kbank_pdf(pdf_path: str) -> list[TransactionRow]:
    """Parse KBank statement PDF."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 4:
                        continue
                    # KBank format: Date | Description | Withdrawal | Deposit | Balance
                    date_str = str(row[0] or "").strip()
                    desc = str(row[1] or "").strip()

                    # Skip header rows
                    if not date_str or "วันที่" in date_str or "Date" in date_str:
                        continue
                    if not re.match(r'\d{2}/\d{2}/\d{2,4}', date_str):
                        continue

                    debit = clean_amount(row[2] if len(row) > 2 else "")
                    credit = clean_amount(row[3] if len(row) > 3 else "")
                    balance = clean_amount(row[4] if len(row) > 4 else "")

                    # Normalize date
                    try:
                        d = datetime.strptime(date_str, "%d/%m/%Y")
                    except:
                        try:
                            d = datetime.strptime(date_str, "%d/%m/%y")
                        except:
                            d = None
                    date_formatted = d.strftime("%Y-%m-%d") if d else date_str

                    rows.append(TransactionRow(
                        date=date_formatted,
                        description=desc,
                        ref="",
                        debit=debit,
                        credit=credit,
                        balance=balance,
                        category=categorize(desc)
                    ))
    return rows


def parse_scb_pdf(pdf_path: str) -> list[TransactionRow]:
    """Parse SCB (Siam Commercial Bank) statement PDF."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 4:
                        continue
                    date_str = str(row[0] or "").strip()
                    desc = str(row[1] or "").strip()

                    if not date_str or "วันที่" in date_str:
                        continue
                    if not re.match(r'\d{2}[-/]\d{2}[-/]\d{2,4}', date_str):
                        continue

                    date_str_clean = date_str.replace("-", "/")
                    try:
                        d = datetime.strptime(date_str_clean, "%d/%m/%Y")
                    except:
                        try:
                            d = datetime.strptime(date_str_clean, "%d/%m/%y")
                        except:
                            d = None
                    date_formatted = d.strftime("%Y-%m-%d") if d else date_str

                    debit = clean_amount(row[2] if len(row) > 2 else "")
                    credit = clean_amount(row[3] if len(row) > 3 else "")
                    balance = clean_amount(row[4] if len(row) > 4 else "")
                    ref = str(row[5] if len(row) > 5 else "").strip()

                    rows.append(TransactionRow(
                        date=date_formatted,
                        description=desc,
                        ref=ref,
                        debit=debit,
                        credit=credit,
                        balance=balance,
                        category=categorize(desc)
                    ))
    return rows


def parse_bbl_pdf(pdf_path: str) -> list[TransactionRow]:
    """Parse Bangkok Bank (BBL) statement PDF."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            # BBL often uses text layout instead of tables
            lines = text.split("\n")
            for line in lines:
                # Pattern: DD/MM/YYYY or DD-MM-YYYY at start
                match = re.match(
                    r'(\d{2}[/-]\d{2}[/-]\d{2,4})\s+(.+?)\s+([\d,]+\.\d{2}|-)\s+([\d,]+\.\d{2}|-)\s+([\d,]+\.\d{2})',
                    line.strip()
                )
                if match:
                    date_str = match.group(1).replace("-", "/")
                    try:
                        d = datetime.strptime(date_str, "%d/%m/%Y")
                    except:
                        try:
                            d = datetime.strptime(date_str, "%d/%m/%y")
                        except:
                            d = None
                    date_formatted = d.strftime("%Y-%m-%d") if d else date_str
                    desc = match.group(2).strip()
                    debit = clean_amount(match.group(3))
                    credit = clean_amount(match.group(4))
                    balance = clean_amount(match.group(5))

                    rows.append(TransactionRow(
                        date=date_formatted,
                        description=desc,
                        ref="",
                        debit=debit,
                        credit=credit,
                        balance=balance,
                        category=categorize(desc)
                    ))
    return rows


def parse_ktb_pdf(pdf_path: str) -> list[TransactionRow]:
    """Parse Krungthai Bank (KTB) statement PDF."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 5:
                        continue
                    date_str = str(row[0] or "").strip()
                    desc = str(row[2] or "").strip()  # KTB has extra ref col

                    if not date_str or not re.match(r'\d{2}[/-]\d{2}[/-]\d{2,4}', date_str):
                        continue

                    date_str_clean = date_str.replace("-", "/")
                    try:
                        d = datetime.strptime(date_str_clean, "%d/%m/%Y")
                    except:
                        try:
                            d = datetime.strptime(date_str_clean, "%d/%m/%y")
                        except:
                            d = None
                    date_formatted = d.strftime("%Y-%m-%d") if d else date_str

                    ref = str(row[1] or "").strip()
                    debit = clean_amount(row[3] if len(row) > 3 else "")
                    credit = clean_amount(row[4] if len(row) > 4 else "")
                    balance = clean_amount(row[5] if len(row) > 5 else "")

                    rows.append(TransactionRow(
                        date=date_formatted,
                        description=desc,
                        ref=ref,
                        debit=debit,
                        credit=credit,
                        balance=balance,
                        category=categorize(desc)
                    ))
    return rows


PARSERS = {
    "kbank": parse_kbank_pdf,
    "scb": parse_scb_pdf,
    "bbl": parse_bbl_pdf,
    "ktb": parse_ktb_pdf,
}


# =====================
# EXCEL GENERATOR
# =====================

BANK_COLORS = {
    "kbank": "1A9F5A",
    "scb": "4E2D8E",
    "bbl": "1E3A8A",
    "ktb": "00A3E0",
}

BANK_NAMES = {
    "kbank": "กสิกรไทย (KBank)",
    "scb": "ไทยพาณิชย์ (SCB)",
    "bbl": "กรุงเทพ (BBL)",
    "ktb": "กรุงไทย (KTB)",
}


def generate_excel(transactions: list[TransactionRow], bank: str, output_path: str, is_pro: bool = False):
    """Generate a formatted Excel file from parsed transactions."""
    wb = openpyxl.Workbook()

    bank_color = BANK_COLORS.get(bank, "333333")
    bank_name = BANK_NAMES.get(bank, bank.upper())

    # ── SHEET 1: Transactions ────────────────────────────────
    ws = wb.active
    ws.title = "รายการเดินบัญชี"

    header_fill = PatternFill("solid", fgColor=bank_color)
    alt_fill = PatternFill("solid", fgColor="F8F9FA")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"📊 รายการเดินบัญชี — ธนาคาร{bank_name}"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=bank_color)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Generated date
    ws["A2"] = f"สร้างเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=10)
    ws["A2"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A2:G2")

    # Header row
    headers = ["วันที่", "รายการ", "อ้างอิง", "เงินออก (บาท)", "เงินเข้า (บาท)", "ยอดคงเหลือ (บาท)"]
    if is_pro:
        headers.append("หมวดหมู่")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[4].height = 28

    # Data rows
    for row_idx, tx in enumerate(transactions, 5):
        is_alt = row_idx % 2 == 0
        row_fill = alt_fill if is_alt else PatternFill("solid", fgColor="FFFFFF")

        data = [tx.date, tx.description, tx.ref, tx.debit or "", tx.credit or "", tx.balance or ""]
        if is_pro:
            data.append(tx.category)

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            # Number formatting
            if col_idx in [4, 5, 6] and value:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

                # Color coding
                if col_idx == 4 and value:  # Debit = red
                    cell.font = Font(color="CC0000")
                elif col_idx == 5 and value:  # Credit = green
                    cell.font = Font(color="006600")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = [14, 45, 20, 18, 18, 20]
    if is_pro:
        col_widths.append(18)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes
    ws.freeze_panes = "A5"

    # Auto-filter
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    # ── SHEET 2: Summary ────────────────────────────────────
    ws2 = wb.create_sheet("สรุปยอด")

    ws2["A1"] = "📈 สรุปยอดรายการ"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:C1")

    total_debit = sum(t.debit for t in transactions)
    total_credit = sum(t.credit for t in transactions)
    tx_count = len(transactions)

    summary_data = [
        ("", "", ""),
        ("รายการทั้งหมด", tx_count, "รายการ"),
        ("เงินออกรวม", total_debit, "บาท"),
        ("เงินเข้ารวม", total_credit, "บาท"),
        ("ยอดสุทธิ (เข้า - ออก)", total_credit - total_debit, "บาท"),
    ]

    for r_idx, (label, val, unit) in enumerate(summary_data, 3):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True if label else False)
        val_cell = ws2.cell(row=r_idx, column=2, value=val)
        if isinstance(val, float):
            val_cell.number_format = '#,##0.00'
            if "ออก" in label:
                val_cell.font = Font(color="CC0000", bold=True)
            elif "เข้า" in label or "สุทธิ" in label:
                val_cell.font = Font(color="006600", bold=True)
        ws2.cell(row=r_idx, column=3, value=unit).font = Font(color="888888")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10

    wb.save(output_path)


def cleanup_files(*paths):
    """Delete temp files after delay."""
    time.sleep(3600)  # Keep for 1 hour then delete
    for p in paths:
        try:
            os.remove(p)
        except:
            pass


# =====================
# ENDPOINTS
# =====================

@app.get("/")
def root():
    return {"status": "ok", "service": "StatementExcel API v1.0"}


@app.get("/health")
def health():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


@app.post("/convert")
async def convert_pdf(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    bank: str = Form(...),
    is_pro: bool = Form(False),
):
    """Convert a bank statement PDF to Excel."""

    # Validate bank
    if bank not in PARSERS:
        raise HTTPException(400, f"ธนาคารไม่รองรับ: {bank}. รองรับ: {', '.join(PARSERS.keys())}")

    # Validate file
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "รองรับเฉพาะไฟล์ .pdf เท่านั้น")

    max_size = 20 * 1024 * 1024  # 20MB
    content = await file.read()
    if len(content) > max_size:
        raise HTTPException(400, "ไฟล์ใหญ่เกิน 20MB")

    # Save upload
    job_id = str(uuid.uuid4())
    upload_path = str(UPLOAD_DIR / f"{job_id}.pdf")
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")

    with open(upload_path, "wb") as f:
        f.write(content)

    try:
        start_time = time.time()

        # Parse PDF
        parser = PARSERS[bank]
        transactions = parser(upload_path)

        if not transactions:
            raise HTTPException(422, "ไม่พบข้อมูลรายการในไฟล์นี้ กรุณาตรวจสอบว่าเป็น statement ที่ถูกต้อง")

        # Generate Excel
        generate_excel(transactions, bank, output_path, is_pro=is_pro)

        elapsed = round(time.time() - start_time, 2)
        total_debit = sum(t.debit for t in transactions)
        total_credit = sum(t.credit for t in transactions)

        # Schedule cleanup
        background_tasks.add_task(cleanup_files, upload_path, output_path)

        # Return file
        filename = f"statement_{bank}_{datetime.now().strftime('%Y%m')}.xlsx"
        return FileResponse(
            output_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "X-Job-Id": job_id,
                "X-Transaction-Count": str(len(transactions)),
                "X-Processing-Time": str(elapsed),
                "X-Total-Debit": str(total_debit),
                "X-Total-Credit": str(total_credit),
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"เกิดข้อผิดพลาดในการแปลงไฟล์: {str(e)}")
    finally:
        # Clean up upload immediately (not output, user needs to download it)
        try:
            os.remove(upload_path)
        except:
            pass


@app.get("/banks")
def get_supported_banks():
    return {
        "banks": [
            {"id": "kbank", "name": "กสิกรไทย", "short": "KBank", "color": "#1A9F5A"},
            {"id": "scb", "name": "ไทยพาณิชย์", "short": "SCB", "color": "#4E2D8E"},
            {"id": "bbl", "name": "กรุงเทพ", "short": "BBL", "color": "#1E3A8A"},
            {"id": "ktb", "name": "กรุงไทย", "short": "KTB", "color": "#00A3E0"},
        ]
    }


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
